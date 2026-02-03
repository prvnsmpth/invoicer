from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
import openpyxl
from openpyxl.utils import get_column_letter

from config import REPORTS_DIR, REPORT_TEMPLATE_PATH


def generate_work_report(
    cycle: Dict,
    events: List[Dict],
    profile: Dict,
    hourly_rate: float,
    output_path: Optional[str] = None
) -> str:
    """Generate Excel work report from invoice cycle data.

    Args:
        cycle: Invoice cycle data
        events: List of calendar events assigned to the cycle
        profile: User profile data
        hourly_rate: Hourly rate in INR
        output_path: Optional custom output path

    Returns:
        Path to the generated report file
    """
    if not REPORT_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Report template not found: {REPORT_TEMPLATE_PATH}")

    # Load template
    wb = openpyxl.load_workbook(REPORT_TEMPLATE_PATH)
    ws = wb.active

    # Group events by title and aggregate hours
    grouped_events = defaultdict(float)
    for event in events:
        grouped_events[event['title']] += event['duration_hours']

    # Sort by title for consistent output
    sorted_items = sorted(grouped_events.items())

    # Fill in name (B2)
    ws['B2'] = profile['full_name']

    # Fill in date (B5) - use cycle end date as the report date
    cycle_end_dt = datetime.strptime(cycle['end_date'], '%Y-%m-%d')
    ws['B5'] = cycle_end_dt

    # Template structure:
    # Row 7: Headers (SR. NO., PROJECT DETAILS, HOURS, REMARKS)
    # Rows 8-12: Empty line item slots (5 rows)
    # Row 13: Empty
    # Row 14: TOTAL HOURS
    # Row 15: RATE / HOUR
    # Row 16: TOTAL (formula)

    FIRST_ITEM_ROW = 8
    TEMPLATE_ITEM_ROWS = 5  # Rows 8-12
    SUMMARY_START_ROW = 14  # Where TOTAL HOURS starts

    num_items = len(sorted_items)
    extra_rows_needed = max(0, num_items - TEMPLATE_ITEM_ROWS)

    # Insert extra rows if needed
    if extra_rows_needed > 0:
        # Insert rows before the summary section
        ws.insert_rows(FIRST_ITEM_ROW + TEMPLATE_ITEM_ROWS, extra_rows_needed)

        # Copy merged cell format from existing item rows for new rows
        for i in range(extra_rows_needed):
            new_row = FIRST_ITEM_ROW + TEMPLATE_ITEM_ROWS + i
            # Merge C:E for project details (same as other item rows)
            ws.merge_cells(f'C{new_row}:E{new_row}')

    # Fill in line items
    total_hours = 0
    for idx, (title, hours) in enumerate(sorted_items):
        row = FIRST_ITEM_ROW + idx
        ws.cell(row=row, column=2, value=idx + 1)  # SR. NO. (column B)
        ws.cell(row=row, column=3, value=title)    # PROJECT DETAILS (column C)
        ws.cell(row=row, column=6, value=hours)    # HOURS (column F)
        total_hours += hours

    # Calculate actual summary row positions
    summary_row = SUMMARY_START_ROW + extra_rows_needed

    # Fill in summary section
    ws.cell(row=summary_row, column=6, value=total_hours)      # TOTAL HOURS value (F14)
    ws.cell(row=summary_row + 1, column=6, value=hourly_rate)  # RATE / HOUR value (F15)

    # Update the formula for TOTAL to reference correct cells
    total_formula_row = summary_row + 2
    ws.cell(row=total_formula_row, column=6, value=f'=F{summary_row}*F{summary_row + 1}')

    # Determine output filename
    if output_path:
        report_path = Path(output_path)
        if not report_path.is_absolute():
            report_path = REPORTS_DIR / report_path
    else:
        # Default: "Work Report (<cycle name>).xlsx"
        report_filename = f"Work Report ({cycle['name']}).xlsx"
        report_path = REPORTS_DIR / report_filename

    # Ensure parent directory exists
    report_path.parent.mkdir(parents=True, exist_ok=True)

    # Save the workbook
    wb.save(report_path)

    return str(report_path)
